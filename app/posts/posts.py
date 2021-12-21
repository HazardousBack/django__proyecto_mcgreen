from django.http.response import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.contrib import messages
from .. import models
from django.db import connection
from ..forms import formulario_cliente, formulario_proveedor, Formulario_registro
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import BORDER_THIN, Border, Side
def registra_usuario(request):
    if request.session.get('email'):
        cursor = connection.cursor()
        if request.method == 'POST':
            formulario = Formulario_registro(request.POST)
            if formulario.is_valid():
                try:
                    ext_email = formulario["slemail"].value()
                    cursor.callproc("AGREGAR_USUARIO",[formulario["matricula"].value(),formulario["nombre_usuario"].value(),formulario["ap_p"].value(),formulario["ap_m"].value(),formulario["sl_puestos"].value(),formulario["email"].value()+ext_email,formulario["contra"].value(),formulario["rol"].value()])
                    mensaje = cursor.fetchall()[0][0]
                    if mensaje == 'USUARIO CREADO':
                        crear = models.Account(username=formulario["email"].value()+formulario["slemail"].value(), email=formulario["email"].value()+formulario["slemail"].value(), password=formulario["contra"].value(), nombre=formulario["nombre_usuario"].value(), ap_paterno=formulario["ap_p"].value(),ap_materno=formulario["ap_m"].value())
                        crear.save()
                finally:
                    cursor.close()
                return redirect("/Registro")
            else:
                print(formulario.errors)
                return HttpResponse("Tu formulario está mal :v")
        return HttpResponse("No hice nada")

def agregar_producto(request):
    if request.session.get('email'):
        if request.method == 'POST':
            email = request.session.get("email")
            try:
                cursor = connection.cursor()
                cursor.callproc("Agrega_INV", [request.POST["producto"], request.POST["descripcion"], request.POST["cantidad"], request.POST["ddw_medidas"], request.POST["ddw_departamentos"], request.POST["precio"], email])
                print(request.POST["producto"]) 
                print(request.POST["descripcion"]) 
                print(request.POST["cantidad"])
                print(request.POST["ddw_medidas"])
                print(request.POST["ddw_departamentos"])
                print(request.POST["precio"]) 
                print(email)
                print(cursor.fetchall()[0])
            finally:
                cursor.close()
            return redirect("/Inventario_general")

def modificar_producto(request):
    if request.method == 'POST':
        cursor = connection.cursor()
        cursor.callproc("MODIFICA_INV", [request.POST["id_producto"], request.POST["producto"], request.POST["descripcion"], request.POST["precio"]])
        if cursor.fetchall()[0][0] != 'EL PRECIO FUE MODIFICADO CORRECTAMENTE':
            messages.error(request, "Ocurrió un error al hacer la modificación")
        messages.success(request, "Elementos editados con éxito")
        return redirect("/Inventario_general")

def descontinuar_producto(request,id_prod):
    if request.session.get("email"):
        cursor = connection.cursor()
        cursor.callproc("DESCONTINUAR_PRODUCTO", [id_prod, request.session.get("email")])
        cursor.close()
        return redirect("/Inventario_general")
    else:
        return redirect("/cerrar_sesion")

def activar_producto(request,id_prod):
    if request.session.get("email"):
        cursor = connection.cursor()
        cursor.callproc("ACTIVAR_PRODUCTO", [id_prod])
        cursor.close()
        return redirect("/Inventario_general")
    else:
        return redirect("/cerrar_sesion")

def generar_compra(request):
    if request.session.get('email'):
        if request.method == 'POST':
            try:
                cursor = connection.cursor()
                cursor.callproc("COMPRA",[request.POST["sl_productos"], request.POST["comprador"], request.POST["compra_id"], request.POST["cantidad"], request.POST["p_u"], request.POST["fecha_compra"], request.POST["sl_proveedores"], request.POST["motivo"]])
            finally:
                cursor.close()
            return redirect("/Compras")
    else:
        return redirect("/cerrar_sesion")

def generar_venta(request):
    if request.session.get('email'):
        if request.method == 'POST':
            cursor = connection.cursor()
            cursor.callproc("VENTA_MOD",[request.POST["sl_productos"].split(' ')[0],request.POST["venta_id"],request.POST["vendedor"],request.POST["cantidad"],request.POST["p_u"],request.POST["fecha"],request.POST["motivo"],request.POST["sl_clientes"],request.POST["articulo"]])
            cursor.close()
            return redirect("/Ventas")

def generar_cuenta_por_cobrar(request):
    if request.session.get('email'):
        if request.method == 'POST':
            cursor = connection.cursor()
            cursor.callproc("VENTA_MOD",[request.POST['email'],request.POST['status'],request.POST['fecha_pago_fac'],request.POST['contrarecibo'],request.POST['fecha_rec_pago'],request.POST['sp'],request.POST['oc'],request.POST['fecha'],request.POST['sl_sistemas'],request.POST['pozo'],request.POST['total_servicios'],request.POST['no_factura'],request.POST['fecha_de_fac'],request.POST['recibo_pago_fac_mcgreen'],request.POST['fecha_r_pag'],request.POST['dolares'],request.POST['monto_mp_pagado']])
            cursor.close()
            return redirect("/Ventas")

def agregar_otros(request):
    if request.method == 'POST':
        cursor = connection.cursor()
        cursor.callproc("MOV_INV", [request.POST["sl_productos"], request.POST["email"], request.POST["id_mov"], request.POST["cantidad"], request.POST["fecha_otro"], request.POST["motivo"], request.POST["sl_tipo_mov"], request.POST["org_des"]])
        cursor.close()
        return redirect("/Otras_E_S")

def conseguir_precio(request,prod):
    if request.is_ajax():
        precios = models.PRECIO_INV_VENTA.objects.filter(ID_PRODUCTO_id=prod).values_list("PRECIO_UNIT_VENTA")[0][0]
        return JsonResponse({'data': precios})
    return HttpResponse("Wrong request")

# Proveedores
def agregar_proveedores(request):
    if request.session.get('email'):
        if request.method == 'POST':
            form = formulario_proveedor(request.POST)
            if form.is_valid():
                cursor = connection.cursor()
                cursor.callproc("Agrega_Proveedor",[form["Identificador"].value(),form["proveedor"].value(),form["telefono"].value(),form["email"].value()])
                cursor.close()
                return redirect("/Compras")    
        return redirect("/Compras")
    else:
        return redirect("/cerrar_sesion")

def agregar_clientes(request):
    if request.session.get('email'):
        if request.method == 'POST':
            form = formulario_cliente(request.POST)
            if form.is_valid():
                cursor = connection.cursor()
                cursor.callproc("Agrega_CLIENTE",[form["Identificador"].value(), form["cliente"].value(), form["direccion"].value(), form["telefono"].value(), form["email"].value()])
                cursor.close()
                return redirect("/Compras")    
        return redirect("/Compras")
    else:
        return redirect("/cerrar_sesion")

def exportar_inventario_excel(request):
    if request.method == 'POST':
        cursor = connection.cursor()
        cursor.callproc("MOSTRAR_PRODUCTOS_ACTIVOS")
        filas = cursor.fetchall()
        cont = 0
        cont = [cont + 1 for row in filas]
        
        bordes = Border(left=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:B4')
        fila_A1 = ws['A1']
        fila_A1.border = bordes

        filas_C1_G4 = ws['C1:G4']
        for fila in filas_C1_G4:
            fila[0].border = bordes

        ws['C1'] = 'PRODUCTOS EXISTENTES'
        ws.merge_cells('C1:G4')
        fila_C1 = ws['C1']
        fila_C1.font = Font(bold=True, color="00008000", size=20, name="Arial")
        fila_C1.alignment = Alignment(horizontal="center", vertical="center")
        fila_C1.border = bordes
        
        ws['H1'] = 'Código:'
        fila_H1 = ws['H1']
        fila_H1.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H1.alignment = Alignment(horizontal="left", vertical="center")
        fila_H1.border = bordes
        fila_I1 = ws['I1']
        fila_I1.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_I1.alignment = Alignment(horizontal="left", vertical="center")
        fila_I1.border = bordes
        
        ws['H2'] = 'Revisión:'
        fila_H2 = ws['H2']
        fila_H2.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H2.alignment = Alignment(horizontal="left", vertical="center")
        fila_H2.border = bordes
        fila_I2 = ws['I2']
        fila_I2.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_I2.alignment = Alignment(horizontal="left", vertical="center")
        fila_I2.border = bordes
        
        filas_H3_H4 = ws['H3:H4']
        for fila in filas_H3_H4:
            fila[0].border = bordes
        ws['H3'] = 'Página:'
        ws.merge_cells('H3:H4')
        fila_H3 = ws['H3']
        fila_H3.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H3.alignment = Alignment(horizontal="left", vertical="center")
        fila_H3.border = bordes
        filas_J3_J4 = ws['I3:J4']
        for fila in filas_J3_J4:
            fila[0].border = bordes
        ws.merge_cells('I3:I4')
        fila_H3 = ws['I3']
        fila_H3.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H3.alignment = Alignment(horizontal="left", vertical="center")
        fila_H3.border = bordes



        ws['A5'] = 'Identificador'
        ws['A5'].font = Font(name="Arial", size=12)
        ws['B5'] = 'Producto'
        ws['B5'].font = Font(name="Arial", size=12)
        ws['C5'] = 'Descripción'
        ws['C5'].font = Font(name="Arial", size=12)
        ws['D5'] = 'Cantidad'       
        ws['D5'].font = Font(name="Arial", size=12)
        ws['E5'] = 'Medida'
        ws['E5'].font = Font(name="Arial", size=12)
        ws['F5'] = 'Departamento'       
        ws['F5'].font = Font(name="Arial", size=12)
        ws['G5'] = 'Precio unitario'
        ws['G5'].font = Font(name="Arial", size=12)
        ws['H5'] = 'Subtotal'       
        ws['H5'].font = Font(name="Arial", size=12)
        ws['I5'] = 'Precio total'       
        ws['I5'].font = Font(name="Arial", size=14)
        filas = 6
        
        for producto in cursor:
            ws.cell(row=filas,column=1).value = producto[0]
            ws.cell(row=filas,column=1).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=1).border = bordes
            ws.cell(row=filas,column=2).value = producto[1]
            ws.cell(row=filas,column=2).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=2).border = bordes
            ws.cell(row=filas,column=3).value = producto[2]
            ws.cell(row=filas,column=3).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=3).border = bordes
            ws.cell(row=filas,column=4).value = producto[3]
            ws.cell(row=filas,column=4).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=4).border = bordes
            ws.cell(row=filas,column=5).value = producto[4]
            ws.cell(row=filas,column=5).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=5).border = bordes
            ws.cell(row=filas,column=6).value = producto[5]
            ws.cell(row=filas,column=6).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=6).border = bordes
            ws.cell(row=filas,column=7).value = producto[6]
            ws.cell(row=filas,column=7).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=7).border = bordes
            ws.cell(row=filas,column=8).value = producto[7]
            ws.cell(row=filas,column=8).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=8).border = bordes
            ws.cell(row=filas,column=9).value = producto[8]
            ws.cell(row=filas,column=9).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=9).border = bordes
            if filas != cont * 2:
                filas = filas + 1

        nombre_archivo ="Inventario.xlsx"
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        cursor.close()
        return response
    else:
        return HttpResponse("<h1>No se pudo descargar el archivo<h1>")

def exportar_compras_excel(request):
    if request.method == 'POST':
        cursor = connection.cursor()
        cursor.callproc("MOSTRAR_COMPRAS")
        filas = cursor.fetchall()
        cont = 0
        cont = [cont + 1 for row in filas]

        bordes = Border(left=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
        
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:B4')
        fila_A1 = ws['A1']
        fila_A1.border = bordes

        filas_C1_G4 = ws['C1:G4']
        for fila in filas_C1_G4:
            fila[0].border = bordes

        ws['C1'] = 'COMPRAS REALIZADAS'
        ws.merge_cells('C1:G4')
        fila_C1 = ws['C1']
        fila_C1.font = Font(bold=True, color="00008000", size=20, name="Arial")
        fila_C1.alignment = Alignment(horizontal="center", vertical="center")
        fila_C1.border = bordes

        fila_B1 = ws['A1']
        fila_B1.font = Font(bold=True, color="00008000", size=20, name="Arial")
        fila_B1.alignment = Alignment(horizontal="center", vertical="center")
        ws['H1'] = 'Código:'
        fila_H1 = ws['H1']
        fila_H1.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H1.alignment = Alignment(horizontal="left", vertical="center")
        fila_H1.border = bordes
        fila_I1 = ws['I1']
        fila_I1.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_I1.alignment = Alignment(horizontal="left", vertical="center")
        fila_I1.border = bordes
        
        ws['H2'] = 'Revisión:'
        fila_H2 = ws['H2']
        fila_H2.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H2.alignment = Alignment(horizontal="left", vertical="center")
        fila_H2.border = bordes
        fila_I2 = ws['I2']
        fila_I2.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_I2.alignment = Alignment(horizontal="left", vertical="center")
        fila_I2.border = bordes
        
        filas_H3_H4 = ws['H3:H4']
        for fila in filas_H3_H4:
            fila[0].border = bordes
        ws['H3'] = 'Página:'
        ws.merge_cells('H3:H4')
        fila_H3 = ws['H3']
        fila_H3.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H3.alignment = Alignment(horizontal="left", vertical="center")
        fila_H3.border = bordes
        filas_J3_J4 = ws['I3:J4']
        for fila in filas_J3_J4:
            fila[0].border = bordes
        ws.merge_cells('I3:I4')
        fila_H3 = ws['I3']
        fila_H3.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H3.alignment = Alignment(horizontal="left", vertical="center")
        fila_H3.border = bordes

        ws['A5'] = 'Id compra'
        ws['A5'].font = Font(name="Arial", size=12)
        ws['B5'] = 'Cod. detalle'
        ws['B5'].font = Font(name="Arial", size=12)
        ws['C5'] = 'Proveedor'
        ws['C5'].font = Font(name="Arial", size=12)
        ws['D5'] = 'Cod. producto'       
        ws['D5'].font = Font(name="Arial", size=12)
        ws['E5'] = 'Producto'
        ws['E5'].font = Font(name="Arial", size=12)
        ws['F5'] = 'Cantidad por producto'       
        ws['F5'].font = Font(name="Arial", size=12)
        ws['G5'] = 'Precio unitario'
        ws['G5'].font = Font(name="Arial", size=12)
        ws['H5'] = 'Medida'       
        ws['H5'].font = Font(name="Arial", size=12)
        ws['I5'] = 'Costo total'       
        ws['I5'].font = Font(name="Arial", size=12)
        filas = 6
        
        for producto in cursor:
            ws.cell(row=filas,column=1).value = producto[0]
            ws.cell(row=filas,column=1).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=1).border = bordes
            ws.cell(row=filas,column=2).value = producto[1]
            ws.cell(row=filas,column=2).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=2).border = bordes
            ws.cell(row=filas,column=3).value = producto[2]
            ws.cell(row=filas,column=3).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=3).border = bordes
            ws.cell(row=filas,column=4).value = producto[3]
            ws.cell(row=filas,column=4).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=4).border = bordes
            ws.cell(row=filas,column=5).value = producto[4]
            ws.cell(row=filas,column=5).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=5).border = bordes
            ws.cell(row=filas,column=6).value = producto[5]
            ws.cell(row=filas,column=6).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=6).border = bordes
            ws.cell(row=filas,column=7).value = producto[6]
            ws.cell(row=filas,column=7).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=7).border = bordes
            ws.cell(row=filas,column=8).value = producto[7]
            ws.cell(row=filas,column=8).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=8).border = bordes
            ws.cell(row=filas,column=9).value = producto[8]
            ws.cell(row=filas,column=9).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=9).border = bordes
            if filas != cont * 2:
                filas = filas + 1

        nombre_archivo ="Compras.xlsx"
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        cursor.close()
        return response
    else:
        return HttpResponse("<h1>No se pudo descargar el archivo<h1>")

def exportar_ventas_excel(request):
    if request.method == 'POST':
        cursor = connection.cursor()
        cursor.callproc("MOSTRAR_VENTAS")
        filas = cursor.fetchall()
        cont = 0
        cont = [cont + 1 for row in filas]
        
        bordes = Border(left=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:B4')
        fila_A1 = ws['A1']
        fila_A1.border = bordes

        filas_C1_G4 = ws['C1:G4']
        for fila in filas_C1_G4:
            fila[0].border = bordes

        ws['C1'] = 'VENTAS REALIZADAS'
        ws.merge_cells('C1:G4')
        fila_C1 = ws['C1']
        fila_C1.font = Font(bold=True, color="00008000", size=20, name="Arial")
        fila_C1.alignment = Alignment(horizontal="center", vertical="center")
        fila_C1.border = bordes

        fila_B1 = ws['A1']
        fila_B1.font = Font(bold=True, color="00008000", size=20, name="Arial")
        fila_B1.alignment = Alignment(horizontal="center", vertical="center")
        ws['H1'] = 'Código:'
        fila_H1 = ws['H1']
        fila_H1.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H1.alignment = Alignment(horizontal="left", vertical="center")
        fila_H1.border = bordes
        fila_I1 = ws['I1']
        fila_I1.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_I1.alignment = Alignment(horizontal="left", vertical="center")
        fila_I1.border = bordes
        
        ws['H2'] = 'Revisión:'
        fila_H2 = ws['H2']
        fila_H2.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H2.alignment = Alignment(horizontal="left", vertical="center")
        fila_H2.border = bordes
        fila_I2 = ws['I2']
        fila_I2.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_I2.alignment = Alignment(horizontal="left", vertical="center")
        fila_I2.border = bordes
        
        filas_H3_H4 = ws['H3:H4']
        for fila in filas_H3_H4:
            fila[0].border = bordes
        ws['H3'] = 'Página:'
        ws.merge_cells('H3:H4')
        fila_H3 = ws['H3']
        fila_H3.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H3.alignment = Alignment(horizontal="left", vertical="center")
        fila_H3.border = bordes
        filas_J3_J4 = ws['I3:J4']
        for fila in filas_J3_J4:
            fila[0].border = bordes
        ws.merge_cells('I3:I4')
        fila_H3 = ws['I3']
        fila_H3.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H3.alignment = Alignment(horizontal="left", vertical="center")
        fila_H3.border = bordes

        ws['A5'] = 'Identificador'
        ws['A5'].font = Font(name="Arial", size=12)
        ws['B5'] = 'Producto'
        ws['B5'].font = Font(name="Arial", size=12)
        ws['C5'] = 'Descripción'
        ws['C5'].font = Font(name="Arial", size=12)
        ws['D5'] = 'Cantidad'       
        ws['D5'].font = Font(name="Arial", size=12)
        ws['E5'] = 'Medida'
        ws['E5'].font = Font(name="Arial", size=12)
        ws['F5'] = 'Departamento'       
        ws['F5'].font = Font(name="Arial", size=12)
        ws['G5'] = 'Precio unitario'
        ws['G5'].font = Font(name="Arial", size=12)
        ws['H5'] = 'Subtotal'       
        ws['H5'].font = Font(name="Arial", size=12)
        ws['I5'] = 'Precio total'       
        ws['I5'].font = Font(name="Arial", size=14)
        filas = 6
        
        for producto in cursor:
            ws.cell(row=filas,column=1).value = producto[0]
            ws.cell(row=filas,column=1).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=1).border = bordes
            ws.cell(row=filas,column=2).value = producto[1]
            ws.cell(row=filas,column=2).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=2).border = bordes
            ws.cell(row=filas,column=3).value = producto[2]
            ws.cell(row=filas,column=3).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=3).border = bordes
            ws.cell(row=filas,column=4).value = producto[3]
            ws.cell(row=filas,column=4).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=4).border = bordes
            ws.cell(row=filas,column=5).value = producto[4]
            ws.cell(row=filas,column=5).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=5).border = bordes
            ws.cell(row=filas,column=6).value = producto[5]
            ws.cell(row=filas,column=6).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=6).border = bordes
            ws.cell(row=filas,column=7).value = producto[6]
            ws.cell(row=filas,column=7).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=7).border = bordes
            ws.cell(row=filas,column=8).value = producto[7]
            ws.cell(row=filas,column=8).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=8).border = bordes
            ws.cell(row=filas,column=9).value = producto[8]
            ws.cell(row=filas,column=9).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=9).border = bordes
            if filas != cont * 2:
                filas = filas + 1

        nombre_archivo ="Ventas.xlsx"
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        cursor.close()
        return response
    else:
        return HttpResponse("<h1>No se pudo descargar el archivo<h1>")

def exportar_diferentes_movimientos_excel(request):
    if request.method == 'POST':
        cursor = connection.cursor()
        cursor.callproc("MOSTRAR_MOV_IND")
        filas = cursor.fetchall()
        cont = 0
        cont = [cont + 1 for row in filas]
        
        bordes = Border(left=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:B4')
        fila_A1 = ws['A1']
        fila_A1.border = bordes

        filas_C1_G4 = ws['C1:G4']
        for fila in filas_C1_G4:
            fila[0].border = bordes

        ws['C1'] = 'MOVIMIENTOS REALIZADOS'
        ws.merge_cells('C1:G4')
        fila_C1 = ws['C1']
        fila_C1.font = Font(bold=True, color="00008000", size=20, name="Arial")
        fila_C1.alignment = Alignment(horizontal="center", vertical="center")
        fila_C1.border = bordes

        fila_B1 = ws['A1']
        fila_B1.font = Font(bold=True, color="00008000", size=20, name="Arial")
        fila_B1.alignment = Alignment(horizontal="center", vertical="center")
        ws['H1'] = 'Código:'
        fila_H1 = ws['H1']
        fila_H1.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H1.alignment = Alignment(horizontal="left", vertical="center")
        fila_H1.border = bordes
        fila_I1 = ws['I1']
        fila_I1.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_I1.alignment = Alignment(horizontal="left", vertical="center")
        fila_I1.border = bordes
        
        ws['H2'] = 'Revisión:'
        fila_H2 = ws['H2']
        fila_H2.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H2.alignment = Alignment(horizontal="left", vertical="center")
        fila_H2.border = bordes
        fila_I2 = ws['I2']
        fila_I2.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_I2.alignment = Alignment(horizontal="left", vertical="center")
        fila_I2.border = bordes
        
        filas_H3_H4 = ws['H3:H4']
        for fila in filas_H3_H4:
            fila[0].border = bordes
        ws['H3'] = 'Página:'
        ws.merge_cells('H3:H4')
        fila_H3 = ws['H3']
        fila_H3.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H3.alignment = Alignment(horizontal="left", vertical="center")
        fila_H3.border = bordes
        filas_J3_J4 = ws['I3:J4']
        for fila in filas_J3_J4:
            fila[0].border = bordes
        ws.merge_cells('I3:I4')
        fila_H3 = ws['I3']
        fila_H3.font = Font(bold=True, color="00000000", size=12, name="Arial")
        fila_H3.alignment = Alignment(horizontal="left", vertical="center")
        fila_H3.border = bordes

        ws['A5'] = 'Id'
        ws['A5'].font = Font(name="Arial", size=12)
        ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B5'] = 'Tipo de movimiento'
        ws['B5'].font = Font(name="Arial", size=12)
        ws['C5'] = 'Nombre de producto'
        ws['C5'].font = Font(name="Arial", size=12)
        ws['D5'] = 'Fecha'       
        ws['D5'].font = Font(name="Arial", size=12)
        ws['E5'] = 'Origen/Destino'
        ws['E5'].font = Font(name="Arial", size=12)
        ws['F5'] = 'Departamento'       
        ws['F5'].font = Font(name="Arial", size=12)
        ws['G5'] = 'Motivo'
        ws['G5'].font = Font(name="Arial", size=12)
        ws['H5'] = 'Cantidad'       
        ws['H5'].font = Font(name="Arial", size=12)
        ws['I5'] = 'Medida'       
        ws['I5'].font = Font(name="Arial", size=14)
        filas = 6
        
        for producto in cursor:
            ws.cell(row=filas,column=1).value = producto[0]
            ws.cell(row=filas,column=1).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=1).border = bordes
            ws.cell(row=filas,column=2).value = producto[1]
            ws.cell(row=filas,column=2).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=2).border = bordes
            ws.cell(row=filas,column=3).value = producto[2]
            ws.cell(row=filas,column=3).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=3).border = bordes
            ws.cell(row=filas,column=4).value = producto[3]
            ws.cell(row=filas,column=4).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=4).border = bordes
            ws.cell(row=filas,column=5).value = producto[4]
            ws.cell(row=filas,column=5).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=5).border = bordes
            ws.cell(row=filas,column=6).value = producto[5]
            ws.cell(row=filas,column=6).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=6).border = bordes
            ws.cell(row=filas,column=7).value = producto[6]
            ws.cell(row=filas,column=7).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=7).border = bordes
            ws.cell(row=filas,column=8).value = producto[7]
            ws.cell(row=filas,column=8).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=8).border = bordes
            ws.cell(row=filas,column=9).value = producto[8]
            ws.cell(row=filas,column=9).font = Font(name="Arial", size=11)
            ws.cell(row=filas,column=9).border = bordes
            if filas != cont * 2:
                filas = filas + 1

        nombre_archivo ="Otros_movimientos.xlsx"
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        cursor.close()
        return response
    else:
        return HttpResponse("<h1>No se pudo descargar el archivo<h1>")
